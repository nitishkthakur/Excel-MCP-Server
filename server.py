"""
Excel MCP Server
================
A FastMCP server that exposes tools for reading and exploring Excel (.xlsx) files.

Tools
-----
- list_sheets     : List all sheet names in an Excel file.
- explore_excel   : Show the dimensions (rows × columns) of every sheet.
- retriever       : Fetch rows from a specific sheet or from all sheets at once.
                    The tool auto-detects the number of columns used.

Usage (stdio transport, default)
---------------------------------
    python server.py

Usage (SSE transport for HTTP clients)
---------------------------------------
    python server.py --transport sse --port 8000

Configuration
-------------
The EXCEL_FILE environment variable (or the ``file_path`` parameter accepted by
each tool) must point to a valid .xlsx file.  When ``file_path`` is omitted the
server falls back to the EXCEL_FILE environment variable, and finally to
``data.xlsx`` in the current working directory.
"""

from __future__ import annotations

import os
from typing import Any

import openpyxl
from fastmcp import FastMCP

# ---------------------------------------------------------------------------
# Server bootstrap
# ---------------------------------------------------------------------------

mcp = FastMCP(
    name="Excel MCP Server",
    instructions=(
        "This server provides tools to read and explore Excel (.xlsx) files. "
        "All tools accept an optional `file_path` parameter. "
        "If omitted, the server uses the EXCEL_FILE environment variable or "
        "falls back to 'data.xlsx' in the current working directory.\n\n"
        "Available tools:\n"
        "  • list_sheets   – list every sheet name in the workbook\n"
        "  • explore_excel – show rows × columns for every sheet\n"
        "  • retriever     – fetch rows from one sheet or all sheets\n"
    ),
)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_DEFAULT_FILE = "data.xlsx"


def _resolve_path(file_path: str | None) -> str:
    """Return the Excel file path to use.

    Priority order:
    1. ``file_path`` argument (if provided and non-empty)
    2. ``EXCEL_FILE`` environment variable
    3. ``data.xlsx`` in the current working directory
    """
    if file_path and file_path.strip():
        return file_path.strip()
    env_path = os.environ.get("EXCEL_FILE", "").strip()
    if env_path:
        return env_path
    return _DEFAULT_FILE


def _load_workbook(file_path: str | None) -> tuple[openpyxl.Workbook, str]:
    """Load an openpyxl workbook and return (workbook, resolved_path).

    Raises
    ------
    FileNotFoundError
        When the resolved path does not exist.
    ValueError
        When the file is not a valid Excel workbook.
    """
    path = _resolve_path(file_path)
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Excel file not found: '{path}'. "
            "Pass the correct path via the `file_path` parameter or set the "
            "EXCEL_FILE environment variable."
        )
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open '{path}' as an Excel workbook: {exc}") from exc
    return wb, path


def _sheet_dimensions(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int]:
    """Return (max_row, max_column) for a worksheet.

    ``openpyxl`` exposes ``ws.max_row`` / ``ws.max_column`` but these can be
    ``None`` for empty sheets.  This helper always returns safe integer values.
    """
    rows = ws.max_row or 0
    cols = ws.max_column or 0
    return rows, cols


def _rows_to_records(ws, n_rows: int | None) -> list[dict[str, Any]]:
    """Convert worksheet rows to a list of dicts keyed by column header.

    The first row is treated as the header row.  If the sheet is empty an
    empty list is returned.

    Parameters
    ----------
    ws:
        An openpyxl worksheet (any mode).
    n_rows:
        Maximum number of *data* rows to return (excluding the header row).
        ``None`` means return all rows.
    """
    rows_iter = ws.iter_rows(values_only=True)

    # Read header row
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []

    # Build column names: fall back to "Col_N" for blank headers
    headers = [
        (str(h) if h is not None else f"Col_{i + 1}")
        for i, h in enumerate(raw_headers)
    ]

    records: list[dict[str, Any]] = []
    for i, row in enumerate(rows_iter):
        if n_rows is not None and i >= n_rows:
            break
        records.append(dict(zip(headers, row)))

    return records


# ---------------------------------------------------------------------------
# MCP tools
# ---------------------------------------------------------------------------


@mcp.tool
def list_sheets(
    file_path: str | None = None,
) -> dict[str, Any]:
    """List every sheet name present in an Excel workbook.

    Parameters
    ----------
    file_path:
        Path to the Excel (.xlsx) file.
        If omitted the server checks the EXCEL_FILE environment variable,
        then falls back to 'data.xlsx' in the current directory.

    Returns
    -------
    A dict with two keys:

    - ``file``: the resolved file path that was opened.
    - ``sheets``: a list of sheet name strings in workbook order.

    Examples
    --------
    >>> list_sheets(file_path="sales_data.xlsx")
    {"file": "sales_data.xlsx", "sheets": ["Jan", "Feb", "Mar", "Summary"]}

    >>> list_sheets()
    {"file": "data.xlsx", "sheets": ["Sheet1"]}
    """
    wb, path = _load_workbook(file_path)
    sheet_names = wb.sheetnames
    wb.close()
    return {"file": path, "sheets": sheet_names}


@mcp.tool
def explore_excel(
    file_path: str | None = None,
) -> dict[str, Any]:
    """Show the dimensions (rows × columns) of every sheet in an Excel workbook.

    The row count includes the header row.  The column count is the number of
    columns that contain at least one non-empty cell (i.e. ``max_column`` as
    reported by openpyxl).

    Parameters
    ----------
    file_path:
        Path to the Excel (.xlsx) file.
        If omitted the server checks the EXCEL_FILE environment variable,
        then falls back to 'data.xlsx' in the current directory.

    Returns
    -------
    A dict with two keys:

    - ``file``: the resolved file path that was opened.
    - ``sheets``: a list of objects, one per sheet, each containing:
        - ``name`` (str): sheet name.
        - ``rows`` (int): total number of rows (including the header).
        - ``columns`` (int): total number of columns with data.
        - ``data_rows`` (int): number of data rows (``rows - 1``, or 0 if empty).

    Examples
    --------
    >>> explore_excel(file_path="sales_data.xlsx")
    {
        "file": "sales_data.xlsx",
        "sheets": [
            {"name": "Jan", "rows": 32, "columns": 5, "data_rows": 31},
            {"name": "Summary", "rows": 5, "columns": 3, "data_rows": 4}
        ]
    }
    """
    wb, path = _load_workbook(file_path)
    sheet_info = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows, cols = _sheet_dimensions(ws)
        sheet_info.append(
            {
                "name": name,
                "rows": rows,
                "columns": cols,
                "data_rows": max(rows - 1, 0),
            }
        )
    wb.close()
    return {"file": path, "sheets": sheet_info}


@mcp.tool
def retriever(
    file_path: str | None = None,
    sheet_name: str | None = None,
    n_rows: int | None = 10,
) -> dict[str, Any]:
    """Fetch rows from one sheet or from all sheets in an Excel workbook.

    The tool **auto-detects the number of columns** in each sheet so you never
    need to specify column indices manually.  The first row of every sheet is
    treated as the column header row; subsequent rows become data records.

    Parameters
    ----------
    file_path:
        Path to the Excel (.xlsx) file.
        If omitted the server checks the EXCEL_FILE environment variable,
        then falls back to 'data.xlsx' in the current directory.
    sheet_name:
        Name of the sheet to read.
        - If provided, only that sheet is returned.
        - If omitted (or ``None``), rows are returned from **every** sheet.
    n_rows:
        Maximum number of *data* rows to return per sheet (default ``10``).
        Set to ``null`` / ``None`` to return all rows.
        The header row is never counted against this limit.

    Returns
    -------
    A dict with two keys:

    - ``file``: the resolved file path that was opened.
    - ``sheets``: a list of objects, one per sheet that was read, each with:
        - ``name`` (str): sheet name.
        - ``columns`` (int): number of columns auto-detected from the header row.
        - ``rows_returned`` (int): number of data rows included in ``data``.
        - ``data`` (list[dict]): rows as dicts keyed by column header.

    Raises
    ------
    KeyError
        If ``sheet_name`` is provided but does not exist in the workbook.

    Examples
    --------
    Fetch the first 5 rows from the "Sales" sheet:

    >>> retriever(file_path="sales_data.xlsx", sheet_name="Sales", n_rows=5)
    {
        "file": "sales_data.xlsx",
        "sheets": [
            {
                "name": "Sales",
                "columns": 4,
                "rows_returned": 5,
                "data": [
                    {"Date": "2024-01-01", "Region": "North", "Product": "Widget", "Amount": 250},
                    ...
                ]
            }
        ]
    }

    Fetch the first 3 rows from every sheet:

    >>> retriever(file_path="sales_data.xlsx", n_rows=3)
    {
        "file": "sales_data.xlsx",
        "sheets": [
            {"name": "Jan", "columns": 5, "rows_returned": 3, "data": [...]},
            {"name": "Feb", "columns": 5, "rows_returned": 3, "data": [...]},
            ...
        ]
    }

    Fetch ALL rows from a sheet (no row limit):

    >>> retriever(file_path="report.xlsx", sheet_name="Data", n_rows=None)
    {
        "file": "report.xlsx",
        "sheets": [{"name": "Data", "columns": 7, "rows_returned": 1500, "data": [...]}]
    }
    """
    wb, path = _load_workbook(file_path)

    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            available = wb.sheetnames
            wb.close()
            raise KeyError(
                f"Sheet '{sheet_name}' not found in '{path}'. "
                f"Available sheets: {available}"
            )
        target_sheets = [sheet_name]
    else:
        target_sheets = wb.sheetnames

    result_sheets = []
    for name in target_sheets:
        ws = wb[name]
        _, cols = _sheet_dimensions(ws)
        records = _rows_to_records(ws, n_rows)
        result_sheets.append(
            {
                "name": name,
                "columns": cols,
                "rows_returned": len(records),
                "data": records,
            }
        )

    wb.close()
    return {"file": path, "sheets": result_sheets}


# ---------------------------------------------------------------------------
# Entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    mcp.run()
