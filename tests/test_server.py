"""
Tests for the Excel MCP Server tools.

Each tool is called directly as a Python function so the tests run without
starting the MCP transport layer.
"""

from __future__ import annotations

import os
import sys
import pytest
import openpyxl

# Make the project root importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import server  # noqa: E402  (import after sys.path manipulation)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def sample_xlsx(tmp_path):
    """Create a temporary .xlsx file with two sheets for testing."""
    wb = openpyxl.Workbook()

    # Sheet 1 – "Sales"
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Date", "Region", "Product", "Amount"])
    ws1.append(["2024-01-01", "North", "Widget", 100])
    ws1.append(["2024-01-02", "South", "Gadget", 200])
    ws1.append(["2024-01-03", "East", "Widget", 150])
    ws1.append(["2024-01-04", "West", "Gadget", 300])

    # Sheet 2 – "Summary"
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Region", "Total"])
    ws2.append(["North", 100])
    ws2.append(["South", 200])

    path = str(tmp_path / "sample.xlsx")
    wb.save(path)
    return path


@pytest.fixture()
def empty_xlsx(tmp_path):
    """Create a temporary .xlsx file that has a single empty sheet."""
    wb = openpyxl.Workbook()
    path = str(tmp_path / "empty.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# list_sheets
# ---------------------------------------------------------------------------

class TestListSheets:
    def test_returns_all_sheet_names(self, sample_xlsx):
        result = server.list_sheets(file_path=sample_xlsx)
        assert result["sheets"] == ["Sales", "Summary"]

    def test_returns_file_path(self, sample_xlsx):
        result = server.list_sheets(file_path=sample_xlsx)
        assert result["file"] == sample_xlsx

    def test_single_sheet(self, empty_xlsx):
        result = server.list_sheets(file_path=empty_xlsx)
        assert len(result["sheets"]) == 1

    def test_file_not_found_raises(self, tmp_path):
        missing = str(tmp_path / "missing.xlsx")
        with pytest.raises(FileNotFoundError, match="Excel file not found"):
            server.list_sheets(file_path=missing)


# ---------------------------------------------------------------------------
# explore_excel
# ---------------------------------------------------------------------------

class TestExploreExcel:
    def test_dimensions_sales_sheet(self, sample_xlsx):
        result = server.explore_excel(file_path=sample_xlsx)
        sales = next(s for s in result["sheets"] if s["name"] == "Sales")
        # 1 header + 4 data rows = 5 rows, 4 columns
        assert sales["rows"] == 5
        assert sales["columns"] == 4
        assert sales["data_rows"] == 4

    def test_dimensions_summary_sheet(self, sample_xlsx):
        result = server.explore_excel(file_path=sample_xlsx)
        summary = next(s for s in result["sheets"] if s["name"] == "Summary")
        # 1 header + 2 data rows = 3 rows, 2 columns
        assert summary["rows"] == 3
        assert summary["columns"] == 2
        assert summary["data_rows"] == 2

    def test_all_sheets_present(self, sample_xlsx):
        result = server.explore_excel(file_path=sample_xlsx)
        names = [s["name"] for s in result["sheets"]]
        assert "Sales" in names
        assert "Summary" in names

    def test_file_not_found_raises(self, tmp_path):
        missing = str(tmp_path / "missing.xlsx")
        with pytest.raises(FileNotFoundError):
            server.explore_excel(file_path=missing)


# ---------------------------------------------------------------------------
# retriever
# ---------------------------------------------------------------------------

class TestRetriever:
    def test_fetch_named_sheet_default_rows(self, sample_xlsx):
        result = server.retriever(file_path=sample_xlsx, sheet_name="Sales")
        sheet = result["sheets"][0]
        assert sheet["name"] == "Sales"
        # Default n_rows=10; sheet has only 4 data rows
        assert sheet["rows_returned"] == 4

    def test_fetch_named_sheet_limited_rows(self, sample_xlsx):
        result = server.retriever(
            file_path=sample_xlsx, sheet_name="Sales", n_rows=2
        )
        sheet = result["sheets"][0]
        assert sheet["rows_returned"] == 2
        assert len(sheet["data"]) == 2

    def test_auto_detects_columns(self, sample_xlsx):
        result = server.retriever(file_path=sample_xlsx, sheet_name="Sales")
        sheet = result["sheets"][0]
        assert sheet["columns"] == 4

    def test_data_keyed_by_header(self, sample_xlsx):
        result = server.retriever(
            file_path=sample_xlsx, sheet_name="Sales", n_rows=1
        )
        first_row = result["sheets"][0]["data"][0]
        assert set(first_row.keys()) == {"Date", "Region", "Product", "Amount"}

    def test_fetch_all_sheets_when_no_name(self, sample_xlsx):
        result = server.retriever(file_path=sample_xlsx, n_rows=2)
        names = [s["name"] for s in result["sheets"]]
        assert "Sales" in names
        assert "Summary" in names

    def test_fetch_all_rows_when_n_rows_none(self, sample_xlsx):
        result = server.retriever(
            file_path=sample_xlsx, sheet_name="Sales", n_rows=None
        )
        sheet = result["sheets"][0]
        assert sheet["rows_returned"] == 4  # all 4 data rows

    def test_unknown_sheet_raises_key_error(self, sample_xlsx):
        with pytest.raises(KeyError, match="not found"):
            server.retriever(file_path=sample_xlsx, sheet_name="DoesNotExist")

    def test_file_not_found_raises(self, tmp_path):
        missing = str(tmp_path / "missing.xlsx")
        with pytest.raises(FileNotFoundError):
            server.retriever(file_path=missing)

    def test_only_one_sheet_returned_when_name_given(self, sample_xlsx):
        result = server.retriever(file_path=sample_xlsx, sheet_name="Summary")
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Summary"

    def test_summary_sheet_columns_auto_detected(self, sample_xlsx):
        result = server.retriever(file_path=sample_xlsx, sheet_name="Summary")
        assert result["sheets"][0]["columns"] == 2


# ---------------------------------------------------------------------------
# _resolve_path helper
# ---------------------------------------------------------------------------

class TestResolvePath:
    def test_explicit_path_wins(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_FILE", "env_file.xlsx")
        result = server._resolve_path(str(tmp_path / "explicit.xlsx"))
        assert result == str(tmp_path / "explicit.xlsx")

    def test_env_var_fallback(self, monkeypatch):
        monkeypatch.setenv("EXCEL_FILE", "/some/env/path.xlsx")
        result = server._resolve_path(None)
        assert result == "/some/env/path.xlsx"

    def test_default_fallback(self, monkeypatch):
        monkeypatch.delenv("EXCEL_FILE", raising=False)
        result = server._resolve_path(None)
        assert result == "data.xlsx"
